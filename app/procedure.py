import os
import uuid
import random
import config
import openpyxl
import datetime
import subprocess
import numpy as np
import pandas as pd
import win32com.client
from sqlalchemy import text
from app.logger import logger
from typing import Optional, List
from datetime import datetime, date
from app.logger_decorator import log_dec
from app.logger_context_manager import time_logger
from openpyxl.styles import PatternFill, Font, Border, Side
from app.srvsap_database import engine as dst_engine, conn as srvsap_pyodbc_conn
from app.directum_database import engine as dir_engine, connection_string as dir_connection
from app.queries import (
    truncate_stage,
    get_users_list,
    insert_task_id,
    insert_mapping,
    select_mapping,
    update_comments,
    delete_from_dst,
    truncate_mapping,
    get_tasks_to_export,
    load_from_stage_to_dst,
    truncate_comments_stage,
    select_from_dst_for_excels
)


class Basic_procedure_methods:

    def __init__(
            self
            ,excel_reports_path
            ,export_register_to_csv_path
            ,without_dir_reports:bool = False
            ,upp_users_codes_list:Optional[List[str]]=None
            ,manual_dir_subject_text:str=None
            ,manual_dir_body_text:str=None
    ):
        if not isinstance(without_dir_reports, bool):
            raise TypeError("without_dir_reports must be a bool")

        if manual_dir_subject_text is not None and not isinstance(manual_dir_subject_text, str):
            raise TypeError("manual_dir_subject_text must be an str or None")

        if manual_dir_body_text is not None and not isinstance(manual_dir_body_text, str):
            raise TypeError("manual_dir_body_text must be an str or None")

        if upp_users_codes_list is not None and not isinstance(upp_users_codes_list, list):
            raise TypeError("upp_users_codes_list must be a list or None")

        if upp_users_codes_list:
            for code in upp_users_codes_list:
                if not isinstance(code, str):
                    raise TypeError(f"Элемент списка upp_users_codes_list должен быть str, а не {type(code).__name__}")

        self.reports_path = excel_reports_path  # директория для excel отчетов
        self.export_register_to_csv_path = export_register_to_csv_path  # директория для первичного экспорта данных 1с->csv
        self.without_dir_reports = without_dir_reports  # признак загрузки без формирования excel отчетов и отправки их в директум ответственным (только загрузка в бд)
        self.upp_users_codes = ",".join(f"'{u}'" for u in upp_users_codes_list) if upp_users_codes_list else None  # коды ответственных для выборочной загрузки (коды обмена упп из справочника физ.лица)
        self.manual_dir_subject_text = manual_dir_subject_text # ручной текст в тему задачи
        self.manual_dir_body_text = manual_dir_body_text # ручной текст в тело задачи


    def exec_query(self, engine, query, params=None):
        # метод для выполнения insert, delete, truncate запросов через sqlalchemy
        if params is None:
            params = {}
        try:
            with engine.connect() as con:
                result = con.execute(text(query), params)
                con.commit()
            return

        except Exception as e:
            con.rollback()
            raise

    def fetch_query(self, engine, query):
        # метод для выполнения select запросов через sqlalchemy, возвращает готовый df
        try:
            with engine.connect() as con:
                result = con.execute(text(query))
                rows = result.fetchall()
                columns = result.keys()
                df = pd.DataFrame(rows, columns=columns)
            return df

        except Exception as e:
            raise

    def convert_to_uuid(self, value):
        # метод конвертирует bytes в привычный uuid
        if isinstance(value, bytes):  # Если значение в формате bytes
            try:
                return str(uuid.UUID(bytes=value))  # Преобразуем в строку UUID
            except Exception as e:
                return None  # Если ошибка, возвращаем None
        return value  # Если это уже строка, оставляем как есть

    @log_dec
    def create_clean_dir(self, path):
        # метод создает (при необходимости) и очищает директорию

        # Создаем директорию, если она не существует
        if not os.path.exists(path):
            os.makedirs(path)

        # в это множество запишем файлы которые не удалось удалить
        self.files_na_to_del = set()

        # очищаем директорию
        for filename in os.listdir(path):
            file_path = os.path.join(path, filename)
            if os.path.isfile(file_path):
                try:
                    os.remove(file_path)
                    logger.info(f"file {file_path} deleted from {path}")
                except PermissionError:
                    self.files_na_to_del.add(file_path)
                    logger.warning(
                        f"file {file_path} from {path} directory is being used by another process!")
                except Exception as e:
                    self.files_na_to_del.add(file_path)
                    logger.warning(f"error while deleting {file_path} file from {path} directory: {e}")

    @log_dec
    def get_users_dict(self) -> dict:
        # Метод получает словарь из директума NameAn:Kod
        self.users_df = self.fetch_query(dir_engine, get_users_list)
        users_dict = {name.replace('ё', 'е').replace('Ё', 'Е').replace('Наталия', 'Наталья'): code for name, code in
                      zip(self.users_df["NameAn"], self.users_df["Kod"])}

        return users_dict

    def get_user_from_excel(self, file_name: str) -> str:
        # Метод получает ФИО из названия отчета Excel
        base_name = os.path.splitext(file_name)[0]
        parts = base_name.split("_", 1)
        user_name = parts[0]

        return user_name


class TaskCreator(Basic_procedure_methods):
    '''
    Объект для первичной выгрузки данных из регистра 1С в БД,
    создания excel отчетов и их рассылки ответственным в Direcum
    '''

    def run(self):

        # очищаем (и если надо создаем) директорию reports_path
        self.create_clean_dir(self.reports_path)

        # очищаем (и если надо создаем) директорию export_register_to_csv_path
        self.create_clean_dir(self.export_register_to_csv_path)

        # запускаем запрос delete_from_dst
        with time_logger("delete_from_dst"):
            self.exec_query(dst_engine, delete_from_dst)

        # запускаем запрос truncate_stage
        with time_logger("truncate_stage"):
            self.exec_query(dst_engine, truncate_stage)

        # экспортируем данные из 1c в csv
        self.sqlcmd_export(self.upp_users_codes)

        # готовим csv данные и запрос для импорта
        self.data_to_load, self.query = self.prepare_csv_data_to_import()

        # импортируем из csv в stage
        self.pyodbc_import(self.data_to_load, self.query)

        # переносим в ReceivablePayableBalanceReport данные из ReceivablePayableBalanceReportStage вместе с кодами Directum
        with time_logger("load_from_stage_to_dst"):
            self.exec_query(dst_engine, load_from_stage_to_dst)

        if self.without_dir_reports is False:
            # формируем df из данных из ReceivablePayableBalanceReport c актуальными комментариями
            with time_logger("select_from_dst_with_comments"):
                self.df = self.fetch_query(dst_engine, select_from_dst_for_excels)

            # создаем excel отчеты
            self.create_reports(self.df)

            # преобразовываем excel отчеты
            self.transform_reports(self.reports_path)

            # создаем задачи в Directum, записываем ИД Задачи в dst таблицу
            self.dir_procedure()

    @log_dec
    def sqlcmd_export(self, upp_users_codes:Optional[List[str]]=None):
        try:
            """Экспорт данных в CSV"""
            if upp_users_codes:
                exp_to_csv_sqlcmd_command_with_upp_users_codes = config.exp_to_csv_sqlcmd_command_specific_users
                exp_to_csv_sqlcmd_command_with_upp_users_codes.append("-v")
                exp_to_csv_sqlcmd_command_with_upp_users_codes.append(f"upp_users_codes={upp_users_codes}")
                sqlcmd_export = subprocess.run(exp_to_csv_sqlcmd_command_with_upp_users_codes, shell=True, text=True,
                                               capture_output=True)
            else:
                sqlcmd_export = subprocess.run(config.exp_to_csv_sqlcmd_command, shell=True, text=True,
                                               capture_output=True)

            if sqlcmd_export.returncode == 0:
                logger.info(f"Successful data export from {config.UPP_DB_NAME} to {config.csv_data}")
                if sqlcmd_export.stdout:
                    logger.info(f"STDOUT: %s", sqlcmd_export.stdout)
            else:
                logger.info(f"Unsuccessful data export from {config.UPP_DB_NAME} to {config.csv_data}")
                if sqlcmd_export.stderr:
                    logger.info(f"STDERR: %s", sqlcmd_export.stderr)

        except Exception as e:
            logger.info(f"Unsuccessful data export from {config.UPP_DB_NAME} to {config.csv_data}")
            if sqlcmd_export.stderr:
                logger.info(f"STDERR: %s", sqlcmd_export.stderr)
            raise

    @log_dec
    def prepare_csv_data_to_import(self):
        # готовим csv данные для вставки в stage

        try:

            # Читаем CSV, пропуская строку с прочерками
            df = pd.read_csv(
                config.csv_data,
                sep=",",
                quotechar='"',
                encoding="utf-8",
                engine="python",
                header=0,  # Заголовки колонок
                skipfooter=2,
                skiprows=[1]  # Пропускаем строку с прочерками
            )

            df = df.replace({np.nan: None})

            # Вместо NULL в датафрейм встает пандасовский NaN, заменяем на питоновский None чтобы при вставке получился NULL
            df = df.where(pd.notna(df), None)

            # Преобразование дат
            df["DocDate"] = pd.to_datetime(df["DocDate"], errors="coerce")

            # Преобразуем DataFrame в список кортежей с обычными Python-типами
            data = [tuple(map(str, row)) for row in df.astype(str).itertuples(index=False, name=None)]

            # исключаем "битые" записи, был случай 05.03.2025 в регистре была некорректная запись
            data = [
                row for row in data
                if row[4] is not None and row[4] != "None"  # DocID не должен быть пустым
                   and row[7] is not None and row[7] != "NaT"  # DocDate должен быть нормальным
            ]

            # Вместо текста "None" делаем настоящий None, иначе в бд зайдет не NULL а текст "None"
            data = [
                tuple(None if x == "None" else x for x in row)
                for row in data
            ]

            # SQL-запрос для вставки
            columns = ", ".join(df.columns)
            placeholders = ", ".join(["?" for _ in df.columns])
            query = f"INSERT INTO {config.stage_table_name} ({columns}) VALUES ({placeholders})"

            logger.info(f"Data and query prepared for import to stage")

            return data, query

        except Exception as e:
            logger.info(f"Data and query are not prepared for import to stage, error: {e}")
            raise

    @log_dec
    def pyodbc_import(self, data, query):
        """Импорт данных из CSV в CSSAPDB2 Stage таблицу"""

        with time_logger('insert into stage from csv data'):

            # Подключение к SQL Server
            try:
                conn = srvsap_pyodbc_conn
                conn.autocommit = False
                cursor = conn.cursor()

                # Проверяем поддержку fast_executemany
                if hasattr(cursor, "fast_executemany"):
                    cursor.fast_executemany = True

                batch_size = 1000
                for i in range(0, len(data), batch_size):

                    batch = data[i:i + batch_size]  # Берем текущую порцию данных

                    try:
                        cursor.executemany(query, batch)
                        logger.info(f"{i + len(batch)} rows processed")


                    except Exception as e:
                        logger.info(f"Error in {i} - {i + len(batch)} batch: {e}")

                        # Проверяем каждую строку по отдельности, чтобы найти проблемную
                        for j, row in enumerate(batch):

                            try:
                                cursor.execute(query, row)  # Пробуем вставить построчно

                            except Exception as row_error:
                                logger.info(f"Error in row {i + j}: {row_error}")
                                conn.rollback()
                                break  # Останавливаемся на первой проблемной строке

                        break  # Остановим весь процесс, как только найдем ошибку

                # Фиксируем транзакцию
                conn.commit()
                logger.info(f"Data is successfully inserted: {len(data)} rows")

            except Exception as e:
                logger.info(f"Insert data error: {e}")
                conn.rollback()

            finally:
                cursor.close()
                conn.close()

    @log_dec
    def create_reports(self, df: pd.DataFrame):
        # Метод создает Excel таблицы по ответственным

        detail_columns = [
            "Контрагент",
            "Тип договора",
            "Договор",
            "Документ расчета с контрагентом",
            "<= 10 дней",
            "> 10 дней",
            "Итог"
        ]

        agg_df = (
            df.groupby(["Ответственный", "Контрагент", "Тип договора"], as_index=False)
            .agg({
                "<= 10 дней": "sum",
                "> 10 дней": "sum",
                "Итог": "sum",
                "Сумма взаимозачетов": "sum",
                "Итог с учетом взаимозачетов": "sum",
                "Актуальный комментарий": "last"  # или "max"/"first"
            })
        )

        # исключаем строки, которые после агрегации получают 0 с учетом взаимозачетов
        # agg_df = agg_df[agg_df["Итог с учетом взаимозачетов"] != 0] #правки от 07.05.2025 Острый Е.А. согл. с Артеменковым М.П.
        agg_df = agg_df[agg_df["Итог с учетом взаимозачетов"] > 0]

        # Итоговый порядок столбцов (на первом листе)
        summary_columns = [
            "Контрагент",
            "Тип договора",
            "<= 10 дней",
            "> 10 дней",
            "Итог",
            "Сумма взаимозачетов",
            "Итог с учетом взаимозачетов",
            "Актуальный комментарий"
        ]

        # -----------------------------
        # ДАЛЕЕ СОХРАНЯЕМ ОТЧЁТ ДЛЯ КАЖДОГО ОТВЕТСТВЕННОГО
        # -----------------------------

        current_date = datetime.now().strftime("%Y-%m-%d")

        # Группируем общий df по "Ответственный", чтобы создать ФАЙЛ
        for responsible_name, sub_df in df.groupby("Ответственный"):

            # 1) Создаем workbook "с нуля"
            workbook = openpyxl.Workbook()

            # 2) Первый лист (агрегированный)
            #    Извлекаем из agg_df только строки для этого "Ответственного"
            current_agg = agg_df[agg_df["Ответственный"] == responsible_name].copy()

            # Удаляем тех. колонку "Ответственный"
            if "Ответственный" in current_agg.columns:
                current_agg.drop(columns=["Ответственный"], inplace=True)

            # Добавляем итого-строку
            summary_row = {}
            summary_row["Контрагент"] = "Итог"
            summary_row["Тип договора"] = "Итог"
            summary_row["<= 10 дней"] = current_agg["<= 10 дней"].sum()
            summary_row["> 10 дней"] = current_agg["> 10 дней"].sum()
            summary_row["Итог"] = current_agg["Итог"].sum()
            summary_row["Сумма взаимозачетов"] = current_agg["Сумма взаимозачетов"].sum()
            summary_row["Итог с учетом взаимозачетов"] = current_agg["Итог с учетом взаимозачетов"].sum()
            summary_row["Актуальный комментарий"] = "Итог"

            current_agg = pd.concat(
                [current_agg, pd.DataFrame([summary_row])],
                ignore_index=True
            )

            # Теперь создаём лист с именем = ответственный (не более 31 символа)
            sheet_name_1 = responsible_name[:31] if responsible_name else "Unknown"
            ws_summary = workbook.active
            ws_summary.title = sheet_name_1

            # Приводим колонки к нужному порядку
            current_agg = current_agg[summary_columns]

            # Записываем current_agg на первый лист
            for r_idx, row_data in enumerate(current_agg.itertuples(index=False), start=1):
                if r_idx == 1:

                    # Заголовки
                    for c_idx, col_name in enumerate(current_agg.columns, start=1):
                        ws_summary.cell(row=r_idx, column=c_idx, value=col_name)

                # Данные
                for c_idx, value in enumerate(row_data, start=1):
                    ws_summary.cell(row=r_idx + 1, column=c_idx, value=value)

            # 3) Второй лист (Детализация) — берём sub_df (в нём уже только этот ответственный)
            #    и извлекаем нужные колонки detail_columns

            current_detail = sub_df[detail_columns].copy()
            current_detail = current_detail[current_detail["Итог"] > 0]

            # Добавляем итого-строку
            detail_summary = {}
            detail_summary["Контрагент"] = "Итог"
            detail_summary["Тип договора"] = "Итог"
            detail_summary["Договор"] = "Итог"
            detail_summary["Документ расчета с контрагентом"] = "Итог"
            detail_summary["<= 10 дней"] = current_detail["<= 10 дней"].sum()
            detail_summary["> 10 дней"] = current_detail["> 10 дней"].sum()
            detail_summary["Итог"] = current_detail["Итог"].sum()

            current_detail = pd.concat(
                [current_detail, pd.DataFrame([detail_summary])],
                ignore_index=True
            )

            ws_detail = workbook.create_sheet(title="Детализация")

            # Записываем данные detail
            for r_idx, row_data in enumerate(current_detail.itertuples(index=False), start=1):
                if r_idx == 1:

                    # Заголовки
                    for c_idx, col_name in enumerate(current_detail.columns, start=1):
                        ws_detail.cell(row=r_idx, column=c_idx, value=col_name)

                # Данные
                for c_idx, value in enumerate(row_data, start=1):
                    ws_detail.cell(row=r_idx + 1, column=c_idx, value=value)

            # 4) Сохраняем итоговую книгу
            filename = fr"{self.reports_path}\{responsible_name}_{current_date}.xlsx"
            workbook.save(filename)

    @log_dec
    def transform_reports(self, folder_path: str):
        """
        Метод обрабатывает файлы Excel:
        1. Ищет (по заголовку) столбец "Актуальный комментарий" и,
           если он есть на листе, красит его в желтый и делает заголовок жирным.
        2. Настраивает ширину всех столбцов на основе заголовков и значений.
        3. Добавляет разделители разрядов для чисел в конкретных столбцах.
        4. Фиксирует строку заголовков.
        5. Добавляет границы вокруг ячеек.
        Применяет эти шаги ко всем листам рабочей книги.
        """

        columns_to_separate = [
            "<= 10 дней",
            "> 10 дней",
            "Итог",
            "Сумма взаимозачетов",
            "Итог с учетом взаимозачетов"
        ]

        # Проходим по каждому файлу в папке
        for file_name in os.listdir(folder_path):
            if file_name.endswith(".xlsx") or file_name.endswith(".xls"):
                file_path = os.path.join(folder_path, file_name)
                # Загружаем книгу
                workbook = openpyxl.load_workbook(file_path)
                # Обходим все листы
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    # ------------------------------
                    # 1. Обработка столбца "Актуальный комментарий" (если есть)
                    # ------------------------------
                    comment_col_idx = None
                    for col_idx in range(1, sheet.max_column + 1):
                        cell_value = sheet.cell(row=1, column=col_idx).value
                        if cell_value == "Актуальный комментарий":
                            comment_col_idx = col_idx
                            break
                    if comment_col_idx is not None:
                        # 1a) Делаем заголовок жирным + жёлтым
                        header_cell = sheet.cell(row=1, column=comment_col_idx)
                        header_cell.font = Font(bold=True)
                        header_cell.fill = PatternFill(
                            start_color="FFFF00",
                            end_color="FFFF00",
                            fill_type="solid"
                        )
                        # 1b) Заливаем все ячейки в этом столбце (начиная со второй строки)
                        yellow_fill = PatternFill(
                            start_color="FFFF00",
                            end_color="FFFF00",
                            fill_type="solid"
                        )
                        for row_idx in range(2, sheet.max_row + 1):
                            cell = sheet.cell(row=row_idx, column=comment_col_idx)
                            cell.fill = yellow_fill
                    # ------------------------------
                    # 2. Настройка ширины всех столбцов
                    # ------------------------------
                    for col_idx, col_cells in enumerate(sheet.columns, start=1):
                        max_width = 0
                        for cell in col_cells:
                            if cell.value is not None:
                                if isinstance(cell.value, (int, float)):
                                    display_val = f"{cell.value:,.2f}"
                                else:
                                    display_val = str(cell.value)

                                cell_length = len(display_val)
                                max_width = max(max_width, cell_length)

                        # Добавляем небольшой отступ
                        adjusted_width = max_width + 4
                        sheet.column_dimensions[
                            openpyxl.utils.get_column_letter(col_idx)
                        ].width = adjusted_width
                    # ------------------------------
                    # 3. Добавление разделителей разрядов для чисел
                    # ------------------------------
                    header_row = [cell.value for cell in sheet[1]]  # заголовки (первая строка)
                    # Находим индексы колонок, которые есть в текущем листе
                    column_indices = [
                        header_row.index(col) + 1
                        for col in columns_to_separate
                        if col in header_row
                    ]
                    # Применяем форматирование
                    for col_idx in column_indices:
                        for row in sheet.iter_rows(
                                min_row=2,
                                max_row=sheet.max_row,
                                min_col=col_idx,
                                max_col=col_idx
                        ):
                            for cell in row:
                                if isinstance(cell.value, (int, float)):
                                    cell.number_format = '#,##0.00'
                    # ------------------------------
                    # 4. Фиксация строки заголовков (A2)
                    # ------------------------------
                    sheet.freeze_panes = "A2"
                    # ------------------------------
                    # 5. Добавление границ вокруг ячеек
                    # ------------------------------
                    thin_side = Side(border_style="thin", color="000000")  # чёрный цвет
                    thin_border = Border(
                        top=thin_side,
                        left=thin_side,
                        right=thin_side,
                        bottom=thin_side
                    )
                    for row in sheet.iter_rows(
                            min_row=1,
                            max_row=sheet.max_row,
                            min_col=1,
                            max_col=sheet.max_column
                    ):
                        for cell in row:
                            cell.border = thin_border
                # Сохраняем изменения в файле
                workbook.save(file_path)

    @log_dec
    def create_task(self, dir_conection, responsible_user_code, attachment_direcory):
        # Метод создает задачу и загружает во вложение Excel отчет

        # Подключаемся к COM-объекту SBLogon.LoginPoint
        login_point = win32com.client.Dispatch("SBLogon.LoginPoint")

        directum_app = login_point.GetApplication(dir_conection)

        script = directum_app.ScriptFactory.GetObjectByName("ОтправитьТМАСЗ")

        subject = config.task_subject if not self.manual_dir_subject_text else self.manual_dir_subject_text
        body = config.task_body if not self.manual_dir_body_text else self.manual_dir_body_text

        # Передаём параметры в скрипт
        script.Params.Add("Инициатор", 'Д000086')  # dbo.MBAnalit.Kod Директум текст
        script.Params.Add("Тема", subject)  # текст
        script.Params.Add("Текст", body)  # текст
        script.Params.Add("Ответственный", responsible_user_code)  # dbo.MBAnalit.Kod Директум текст
        script.Params.Add("Документ", attachment_direcory)  # путь к вложению

        result = script.Execute()  # ИД задачи
        return result

    @log_dec
    def dir_procedure(self):
        # Метод выполняет часть процесса связанную с directum
        # Получачет код ответственного
        # Создает задачу в directum и прикладывает excel
        # Записывает ИД созданной задачи в ReceivablePayableBalanceReport

        self.users_dict = self.get_users_dict()

        with time_logger("truncate_mapping"):
            self.exec_query(dst_engine, truncate_mapping)

        for file_name in os.listdir(self.reports_path):

            if file_name.lower().endswith('.xlsx'):
                file_path = os.path.join(self.reports_path, file_name)

                user_name = self.get_user_from_excel(file_name)
                user_code = self.users_dict.get(
                    user_name.replace('ё', 'е').replace('Ё', 'Е').replace('Наталия', 'Наталья'))

                if user_code is None:
                    logger.warning(f"Code for {user_name} not found")
                    continue

                task_id = self.create_task(dir_connection, user_code, file_path)

                with time_logger("insert_task_id"):
                    self.exec_query(dst_engine, insert_task_id, {"task_id": task_id, "user_code": user_code})

                with time_logger("insert_mapping"):
                    self.exec_query(dst_engine, insert_mapping, {"task_id": task_id, "excel_report_name": file_name})


class TaskProcessor(Basic_procedure_methods):
    '''
    Объект для извлечения данных из задач Directum и обновления БД комменатриями из excel отчетов
    '''

    def run(self):

        # очищаем (и если надо создаем) директорию reports_path
        self.create_clean_dir(self.reports_path)

        # получаем из dst таблицы task_id, по которым нужно собрать заполненные excel вложения,
        # собираем excel вложения в reports_path
        self.export_reports()

        # удаляем некорректные или лишние вложения
        self.delete_invalid_reports()

        # собираем комментарии в self.df_to_update_comments
        self.collect_comments()

        # вставляем комментарии в dst
        self.insert_comments()

    @log_dec
    def export_reports(self):

        # метод извлекает из ReceivablePayableBalanceReport все task_id за максимальную дату created_at
        with time_logger('get_tasks_to_export'):
            self.tasks_df = self.fetch_query(dst_engine, get_tasks_to_export)

        # метод по task_id извлекает вложения из Directum и сохраняет в директорию self.reports_path
        for task_id in self.tasks_df['task_id']:
            with time_logger(f"export report from task {task_id}"):
                self.export_documents_from_task(task_id)

    @log_dec
    def export_documents_from_task(self, task_id):
        # метод по task_id извлекает вложения из Directum и сохраняет в директорию self.reports_path

        # Подключаемся к COM-объекту SBLogon.LoginPoint
        login_point = win32com.client.Dispatch("SBLogon.LoginPoint")

        directum_app = login_point.GetApplication(dir_connection)

        script = directum_app.ScriptFactory.GetObjectByName("ExportDocumentsFromTask")

        # Передаём параметры в скрипт
        script.Params.Add("TaskID", task_id)  # Идентификатор задачи
        script.Params.Add("Path", self.reports_path)  # Путь для сохранения документов

        result = script.Execute()

    @log_dec
    def delete_invalid_reports(self):
        # метод удаляет файлы, если их нет в таблице ReceivablePayableBalanceReportTasksExcelsMapping
        # нарпимер, если ответственный добавил еще одно вложение, или переименовал вложение

        # получаем список корректных вложений
        with time_logger('select_mapping'):
            self.valid_reports_df = self.fetch_query(dst_engine, select_mapping)

        # делаем множество
        valid_excel_names = set(self.valid_reports_df["excel_report_name"].dropna())

        # очищаем директорию от лишних вложений
        for filename in os.listdir(self.reports_path):
            file_path = os.path.join(self.reports_path, filename)

            # если формат файла не xlsx, то удаляем
            if not filename.lower().endswith(".xlsx"):
                os.remove(file_path)
                continue

            # удаляем все после первого .xlsx, пока только в переменной new_filename
            xlsx_idx = filename.lower().index(".xlsx")
            new_filename = filename[:xlsx_idx + 5]

            # если filename нет в valid_excel_names, то удаляем
            if new_filename not in valid_excel_names:
                os.remove(file_path)
                continue

            # пытаемся сохранить, если ошибка - удаляем
            new_file_path = os.path.join(self.reports_path, new_filename)

            try:
                os.rename(file_path, new_file_path)

            # если файл с таким именем уже есть - удаляем
            except FileExistsError:
                os.remove(file_path)

            # если другая ошибка - выводим ошибку в лог и удаляем
            except Exception as e:
                logger.info(f"error trying to export {filename} file")
                os.remove(file_path)

    @log_dec
    def collect_comments(self):
        """
        1) Метод обходит все Excel-файлы в директории self.reports_path.
        2) Для каждого файла (первый лист):
           - получает user_name через self.get_user_from_excel(file_name)
           - читает колонки "Контрагент" и "Актуальный комментарий"
           - формирует список строк (user_name, partner_name, comment)
        3) Складывает все строки в общий DataFrame self.df_to_update_comments
        """
        all_rows = []

        # формируем актуальный словарь с ФИО и кодами Directum
        self.users_dict = self.get_users_dict()

        # цикл по каждому файлу в директории
        for file_name in os.listdir(self.reports_path):
            if file_name.lower().endswith(".xlsx"):
                # извлекаем имя ответственного из имени файла
                user_name = self.get_user_from_excel(file_name)

                # получаем Directum код по имени ответственного
                user_code = self.users_dict.get(
                    user_name.replace('ё', 'е').replace('Ё', 'Е').replace('Наталия', 'Наталья'))

                if user_code is None:
                    logger.warning(f"Code for {user_name} not found")
                    continue

                file_path = os.path.join(self.reports_path, file_name)

                # Читаем только ПЕРВЫЙ лист (sheet_name=0)
                df_excel = pd.read_excel(file_path, sheet_name=0)

                # # Проверка, что файл содержит колонки "Контрагент" и "Актуальный комментарий" на первом листе
                # if "Контрагент" not in df_excel.columns or "Актуальный комментарий" not in df_excel.columns:
                #     # Если их нет, пропускаем файл
                #     logger.warning(f"File {file_name} doesn't contain required columns")
                #     continue
                # ДОРАБОТКА  04.06.2025 - добавление ContractType
                # Проверка, что файл содержит колонки "Контрагент", "Тип договора" и "Актуальный комментарий" на первом листе
                if "Контрагент" not in df_excel.columns or "Актуальный комментарий" not in df_excel.columns or "Тип договора" not in df_excel.columns:
                    # Если их нет, пропускаем файл
                    logger.warning(f"File {file_name} doesn't contain required columns")
                    continue

                # ДОРАБОТКА 13.04.2025 ПОСЛЕ ОШИБКИ 10.04.2025
                # EXCEL ЧАЛИКОВОЙ ДАРЬИ СЧИТЫВАЛСЯ НЕКОРРЕКТНО - С ПУСТЫМИ СТРОКАМИ ПОСЛЕ СТРОКИ ИТОГ И ОДНОЙ ПУСТОЙ КОЛОНКОЙ В КОНЦЕ
                # Фильтруем строки: оставляем только строки до первой строки, где в "Контрагент" записано "Итог"
                if "Итог" in df_excel["Контрагент"].values:
                    # Определяем индекс первой строки, где Контрагент == "Итог"
                    idx = df_excel.index[df_excel["Контрагент"] == "Итог"][0]
                    df_excel = df_excel.loc[:idx - 1]  # оставляем строки до этой (не включая её)

                # Отбрасываем лишние столбцы: оставляем столбцы от начала до "Актуальный комментарий"
                col_list = list(df_excel.columns)
                if "Актуальный комментарий" in col_list:
                    col_index = col_list.index("Актуальный комментарий")
                    df_excel = df_excel.iloc[:, :col_index + 1]

                # Фильтруем ненужные строки
                df_excel = df_excel.loc[
                    (df_excel["Контрагент"] != "Итог")  # & (прописать дополнительные условия)
                ]

                # Формируем список кортежей: (user_code, partner_name, comment)
                for idx, row in df_excel.iterrows():
                    partner_name = row["Контрагент"]
                    # ДОРАБОТКА  04.06.2025 - добавление ContractType
                    contract_type = row["Тип договора"]
                    comment = row["Актуальный комментарий"]

                    all_rows.append((user_code, partner_name, contract_type, comment))

        # Создаём единый DataFrame
        self.df_to_update_comments = pd.DataFrame(all_rows,
                                                  columns=["UserCode", "PartnerName", "ContractType", "comment"])
        # если не заполнено поле "Актуальный комментарий" то вместо nan ставим None чтобы в дальнейшем получился NULL
        self.df_to_update_comments["comment"] = self.df_to_update_comments["comment"].replace({np.nan: None})

        # ДОРАБОТКА 13.04.2025 ПОСЛЕ ОШИБКИ 10.04.2025
        # EXCEL ЧАЛИКОВОЙ ДАРЬИ СЧИТЫВАЛСЯ НЕКОРРЕКТНО - С ПУСТЫМИ СТРОКАМИ ПОСЛЕ СТРОКИ ИТОГ И ОДНОЙ ПУСТОЙ КОЛОНКОЙ В КОНЦЕ
        # еще раз убедимся, что в поле Контрагент нет nan значений
        self.df_to_update_comments = self.df_to_update_comments[self.df_to_update_comments["PartnerName"].notna()]

    @log_dec
    def insert_comments(self):
        # Метод очищает и заполняет таблицу ReceivablePayableBalanceReportCommentsStage,
        # а затем комментариями из этой таблицы обновляет комментарии в ReceivablePayableBalanceReport

        with time_logger('truncate comments stage'):
            self.exec_query(dst_engine, truncate_comments_stage)

        with time_logger("insert into ReceivablePayableBalanceReportCommentsStage"):
            # Вставляем данные в таблицу CommentsStage
            with dst_engine.connect() as dst_connection:
                try:
                    BATCH_SIZE = 1000
                    for i in range(0, len(self.df_to_update_comments), BATCH_SIZE):
                        batch = self.df_to_update_comments.iloc[i:i + BATCH_SIZE]
                        batch.to_sql('ReceivablePayableBalanceReportCommentsStage', con=dst_connection,
                                     if_exists='append',
                                     index=False)

                    dst_connection.commit()

                except Exception as e:
                    dst_connection.rollback()
                    raise

        with time_logger('update comments'):
            self.exec_query(dst_engine, update_comments)
